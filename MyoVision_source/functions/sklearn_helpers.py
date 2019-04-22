import scipy as sp
import pandas as pd
import openpyxl
import copy


# START load_fiber_detect_training_set_from_file*******************************
def load_fiber_detect_training_set_from_file(filename, target_col, ignore_cols=None):

    data = sp.genfromtxt(filename, skip_header=1)
    target = data[:, target_col]

    ignore_cols.append(target_col)
    ignore_cols.sort(reverse=True)
    for col in ignore_cols:
        data = sp.delete(data, col, 1)

    return data, target
# END load_fiber_detect_training_set_from_file/////////////////////////////////


# START load_fiber_detect_training_set_from_array******************************
def load_fiber_detect_training_set_from_array(all_data, target_col, ignore_cols=None):
    data = copy.deepcopy(all_data)
    ignore = copy.deepcopy(ignore_cols)
    # Converts Pandas DataFrame to numpy array if possible
    try:
        target = data[:, target_col]
    except TypeError:
        data = data.to_numpy()
        target = data[:, target_col]

    if ignore is None:
        ignore = []
    ignore.append(target_col)
    ignore.sort(reverse=True)
    # ignore_cols is an array that includes the columns numbers to be removed,
    # leaving only the parameter data left
    for col in ignore:
        data = sp.delete(data, col, 1)

    return data, target
# END load_fiber_detect_training_set_from_array////////////////////////////////


def get_num_columns_from_excel(filename):
    wb = openpyxl.load_workbook(filename=filename, read_only=True)
    ws = wb.active
    return ws.max_column


# START load_fiber_detect_training_set_from_excel******************************
def load_fiber_detect_training_set_from_excel(filename, target_col, ignore_cols=None):
    wb = openpyxl.load_workbook(filename=filename, read_only=True)
    ws = wb.active

    def parse_blob_num(str):
        num = str
        if "_" in str:
            num = str.split("_")
            num = num[1]
        return num

    blob_nums = []
    for row in ws.values:
        if row[1] is not None:
            try:
                blob_nums.append(int(parse_blob_num(row[1])))
            except ValueError:
                print("NaN")

    blob_nums.sort()
    print(blob_nums)
    print(len(blob_nums))
        # print(row_cnt)
        # row_cnt += 1
        # col_cnt = 1
        # for value in row:
        #     if value is not None:
        #         if col_cnt is 2:
        #             value = parse_blob_num(value)
        #         print("\t", value)
        #         if "Merged" in value:
        #             print("\tCLASSIFY AS CONNECTED")
        #     col_cnt += 1

