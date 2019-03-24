from functions.display_mult_ims import display_mult_ims, blob_selector
import functions.image_processing as fip
from functions.lists2txt import lists2txt
import matplotlib.pyplot as plt
from dataStructure import *
from PIL import Image, ImageDraw, ImageFont
import math
import sys
import scipy as sp
import skimage.color
import skimage.exposure
import skimage.feature
import skimage.filters
import skimage.measure
import skimage.morphology
import skimage.segmentation
import skimage.util
import cv2 as cv
import time
import copy
import pandas as pd
import xlsxwriter
import openpyxl.utils.dataframe
import openpyxl.drawing.image
import zipfile
import os


# Muscle Stain class:
#     This is a class that describes a membrane stain and aids in its processing
class MuscleStain:
    def __init__(self, mem_im):
        self.name = ''
        self.num_blobs = 0
        self.membrane_image = mem_im
        self.fibers_binary = sp.empty(mem_im[:, :, 0].shape)
        self.edge_im = sp.empty(mem_im[:, :, 0].shape)
        self.blob_labels = sp.empty(mem_im[:, :, 0].shape)
        self.blob_metric_headers = sp.array(["blob_num", "area", "convexity", "eccentricity", "solidity"])
        self.blob_metrics = dataStructure()
        # self.blob_metrics2 = pd.DataFrame

        # Progress Flags
        self.flag_detect = False
        self.flag_properties_1 = False
        self.flag_classify_1 = False
        self.flag_masks_1 = False
        self.flag_separate = False
        self.flag_properties_2 = False
        self.flag_classify_2 = False
        self.flag_masks_2 = False
        self.progress_list = [self.flag_detect,
                              self.flag_properties_1,
                              self.flag_classify_1,
                              self.flag_masks_1,
                              self.flag_separate,
                              self.flag_properties_2,
                              self.flag_classify_2,
                              self.flag_masks_2]


    # START initialize_excel_data**************************************************
    def initialize_excel_data(self, filename):
        if '.xlsx' not in filename:
            filename = filename + '.xlsx'
        self.save_path = os.path.dirname(filename)
        self.excel_book = filename
        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        wb.remove(ws)
        wb.create_sheet('Initial Blob Data')
        wb.create_sheet('Initial Blob Image')
        wb.create_sheet('Initial Blob Classifications')
        wb.create_sheet('Post-Watershed Data')
        wb.create_sheet('Post-Watershed Image')
        wb.create_sheet('Post-Watershed Classifications')
        wb.create_sheet('Contours')

        self.image_sheet_names = ['Initial Blob Image',
                                  'Initial Blob Classifications',
                                  'Post-Watershed Image',
                                  'Post-Watershed Classifications',
                                  'Contours']

        wb.save(filename)

    # END initialize_excel_data////////////////////////////////////////////////////

    # START find_fibers************************************************************
    # Description:
    #     Finds the edges/boundaries of muscle fibers in the image using a
    #     combination of a Frangi filter to get an edge probability map combined with a
    #     k-means clustering algorithm to separate the membrane from the background
    # Inputs:
    #     orig_input_im: RGB or RGBA image specified by an ndarray
    # Returns:
    #     binary_im: Binary image highlighting the muscle fiber membrane
    def find_fibers(self):
        # Converts RGB image to grayscale for processing
        input_im = skimage.color.rgb2gray(self.membrane_image)
        # Performs a 3x3 (default) median filter on the image to reduce noise
        input_im = sp.signal.medfilt(input_im, [3, 3])
        # Invert grayscale image
        input_im = 1 - input_im
        # Normalize 0-1 scale
        input_im = input_im - sp.amin(input_im)
        input_im = input_im / sp.amax(input_im)
        # Contrast Adjustment
        p2, p98 = sp.percentile(input_im, (1, 99))
        input_im = skimage.exposure.rescale_intensity(input_im, in_range=(p2, p98))
        input_im *= 255.0  # Image is already normalized between 0-1

        # Frangi filter: scale range = sigma, beta1 = alpha, beta2 = beta
        fran = skimage.filters.frangi
        edge_im = self.block_proc(input_im, [500, 500], fran)
        frangi_kwargs = {
            'scale_range': (int(1), int(2)),
            'scale_step': 0.25,
            'beta1': 1,
            'beta2': 10}
        frangi_args = {
            (1, 2),
            0.25,
            1,
            10
        }
        additional_kwargs = {
            'dtype': int
        }

        # def wrapped_frangi(array):
        #     print(array.shape)
        #     return skimage.filters.frangi(array,
        #                                   (1, 2),
        #                                   0.25,
        #                                   1,
        #                                   10)
        #
        # edge_im = skimage.util.apply_parallel(function=fran,
        #                                       array=input_im,
        #                                       chunks=None,
        #                                       extra_arguments={(1, 2), 0.25, 1, 10})
        # print('did the frangi')
        # edge_im = fran(input_im, scale_range=(1, 2), scale_step=0.25, beta1=1, beta2=10)
        self.edge_im = edge_im

        # OTSUS METHOD ***********************************************
        thresh = skimage.filters.threshold_otsu(edge_im)
        binary_im = edge_im > thresh
        # END OTSUS METHOD *******************************************

        # Create usable binary image and clear blobs touching border
        binary_im = 1 - binary_im
        binary_im = binary_im.astype(sp.bool_)  # Must recast the array as a boolean array
        binary_im = skimage.segmentation.clear_border(binary_im)

        self.fibers_binary = binary_im
        # Advance flags
        self.flag_detect = True
        self.progress_list[0] = True
    # END find_fibers//////////////////////////////////////////////////////////////


    # START find_fiber_properties**************************************************
    # Description:
    #     Analyzes binary blobs describing muscle fibers and returns properties
    #     relevant to distinguishing between individual muscle fibers, connected
    #     muscle fibers, and interstitial space
    # Inputs:
    #     binary_input_im: binary image where the positive pixels describe blobs
    # Returns:
    #     N/A
    def find_fiber_properties(self, rerun_connected):

        # Checks flags
        if self.flag_detect is False and rerun_connected is False:
            print("Prerequisite find_fibers not run.\n"
                  "Run find_fibers before find_fiber_properties with rerun_connected=False")
            sys.exit(1)
        elif self.flag_separate is False and rerun_connected is True:
            print("Prerequisite separate_connected_fibers not run.\n"
                  "Run separate_connected_fibers before find_fiber_properties with rerun_connected=True")
            sys.exit(1)


        # Remove tiny dots/blobs and fill tiny holes in blobs
        if rerun_connected is False:
            clean_binary_input = skimage.morphology.remove_small_objects(self.fibers_binary)
            clean_binary_input = skimage.morphology.remove_small_holes(clean_binary_input)
        else:
            clean_binary_input = sp.logical_or(self.post_watershed_binary, self.fiber_mask)
            clean_binary_input = sp.logical_or(clean_binary_input, self.space_mask)

        # Label image from fiber binary image
        self.blob_labels, num_blobs = skimage.measure.label(clean_binary_input, return_num=True)
        self.num_blobs = num_blobs
        props = skimage.measure.regionprops(self.blob_labels, coordinates='rc')
        count = 0
        # Arrays for calculated metrics
        # METRICS USED:
        # area
        # convexity
        # extent
        # solidity
        used_metrics = ['Blob Number', 'Area', 'Convexity', 'Extent', 'Eccentricity', 'Solidity']

        self.data = sp.empty([num_blobs, len(used_metrics)])

        self.blob_metrics.area = sp.empty(num_blobs)            # for multiple metrics

        self.blob_metrics.extent = sp.empty(num_blobs)          # extent metric
        self.blob_metrics.perim = sp.empty(num_blobs)           # for convexity calculation
        self.blob_metrics.conv_perim = sp.empty(num_blobs)      # for convexity calculation
        self.blob_metrics.convexity = sp.empty(num_blobs)       # convexity metric
        self.blob_metrics.solidity = sp.empty(num_blobs)        # solidity metric
        self.blob_metrics.eccentricity = sp.empty(num_blobs)    # eccentricity metric

        # Misc calcs
        self.blob_metrics.bboxes = sp.empty([num_blobs, 4])     # bounding boxes
        self.blob_metrics.centroidsx = sp.empty(num_blobs)      # centroid info (x position)
        self.blob_metrics.centroidsy = sp.empty(num_blobs)      # centroid info (y position)

        # Goes through blobs and sets corresponding properties
        for blob in props:
            self.blob_metrics.area[count] = blob.area
            self.blob_metrics.extent[count] = blob.extent
            self.blob_metrics.perim[count] = blob.perimeter
            self.blob_metrics.solidity[count] = blob.solidity
            self.blob_metrics.temp_convex_im = blob.convex_image  # to find individual blob convex hull perimeters
            self.blob_metrics.conv_perim[count] = skimage.measure.perimeter(self.blob_metrics.temp_convex_im)
            self.blob_metrics.convexity[count] = self.blob_metrics.perim[count] / self.blob_metrics.conv_perim[count]

            self.blob_metrics.bboxes[count, :] = blob.bbox
            self.blob_metrics.eccentricity[count] = blob.eccentricity
            self.blob_metrics.centroidsx[count] = blob.centroid[1]
            self.blob_metrics.centroidsy[count] = blob.centroid[0]

            self.data[count, :] = sp.array([count,
                                            blob.area,
                                            self.blob_metrics.convexity[count],
                                            blob.extent,
                                            blob.eccentricity,
                                            blob.solidity])

            count += 1

        max_area = sp.amax(self.data[:, 1]) # column 1 is area column
        self.data[:, 1] = self.data[:, 1]/max_area

        self.output_metrics = pd.DataFrame(self.data, columns=used_metrics)

        # Dumps all numerical data and overlayed image to excel file
        if rerun_connected is False:
            wb = openpyxl.load_workbook(self.excel_book)
            ws_data = wb['Initial Blob Data']
        else:
            wb = openpyxl.load_workbook(self.excel_book)
            ws_data = wb['Post-Watershed Data']
        # Numerical Data to excel
        for r in openpyxl.utils.dataframe.dataframe_to_rows(self.output_metrics, index=False, header=True):
            ws_data.append(r)
        # Image to excel
        ar_im = sp.asarray(clean_binary_input)
        pil_im = fip.numpy_binary_to_pil(ar_im)
        pil_im = self.overlay_blobs(pil_im)
        if rerun_connected is False:
            self.initial_blob_image = pil_im
        else:
            self.post_watershed_image = pil_im

        wb.save(self.excel_book)

        if rerun_connected is False:
            self.flag_properties_1 = True
            self.progress_list[1] = True
        elif rerun_connected is True:
            self.flag_properties_2 = True
            self.progress_list[5] = True
    # END find_fiber_properties////////////////////////////////////////////////////

    # START classify_blobs*********************************************************
    # Description:
    #     Using a given classifier, this function predicts the identity of all
    #     blobs in the blob self.data field
    # Inputs:
    #     classifier: classifier returned from sklearn.SVC
    # Returns:
    #     N/A
    def classify_blobs(self, classifier):

        if self.flag_properties_1 is False:
            print("Prerequisite find_fiber_properties not run.\n"
                  "Run find_fiber_properties before classify_blobs")
            sys.exit(1)
        elif self.flag_classify_1 is True and self.flag_properties_2 is False:
            print("Prerequisite find_fiber_properties not rerun.\n"
                  "Rerun find_fiber_properties before classify_blobs")
            sys.exit(1)

        # classifier is the Support Vector Machine that has been trained.
        # print(self.data[:, 1:].shape)
        self.blob_classified = classifier.predict(self.data[:, 1:])

    # END classify_blobs///////////////////////////////////////////////////////////

    # START get_masks**************************************************************
    # Description:
    #     Creates separate masks for categories of: fibers, connected fibers, space
    # Inputs:
    #     N/A
    # Returns:
    #     N/A
    def get_masks(self):
        blobs = sp.arange(self.data.shape[0])
        # Gets indices where a blob is classified as a fiber
        fiber_idx = blobs[self.blob_classified == 1]
        print(str(len(fiber_idx)) + " fibers found")
        # Gets indices where a blob is classified as a connected fiber
        connected_idx = blobs[self.blob_classified == 2]
        print(str(len(connected_idx)) + " connected fibers found")
        # Gets indices where a blob is classified as a space
        space_idx = blobs[self.blob_classified == 3]
        print(str(len(space_idx)) + " spaces found")

        label_test = self.blob_labels
        # label_height, label_width = label_test.shape
        # print(label_width, "x", label_height, " : ", label_test.shape)

        # flat_labels = sp.ndarray.flatten(label_test)
        print("making fiber mask")
        fiber_mask = sp.isin(label_test, (fiber_idx+1))
        print("making connected mask")
        connected_mask = sp.isin(label_test, (connected_idx+1))
        print("making space mask")
        space_mask = sp.isin(label_test, (space_idx+1))

        self.fiber_mask = fiber_mask
        self.connected_mask = connected_mask
        self.space_mask = space_mask

        all = self.create_classification_image()
        pil_im = fip.numpy_rgb_to_pil(all)
        if self.flag_properties_1 is True and self.flag_properties_2 is False:
            self.mask_image_1 = pil_im
        else:
            self.mask_image_2 = pil_im
    # END get_masks////////////////////////////////////////////////////////////////


    # START separate_connected_fibers**********************************************
    # Description:
    #     Performs watershed transformation on the connected fibers of the image
    #     to try and separate them
    # Inputs:
    #     h_test: (Will be replaced with pixel_size) Used for pre-watershed process
    # Returns:
    #     N/A
    def separate_connected_fibers(self, h_test):
        s = time.time()
        dist = sp.ndimage.distance_transform_edt(self.connected_mask)
        e = time.time()
        print("\tdist transform: ", e-s)
        # dist[1-self.connected_mask] = float('inf')
        dist = -dist
        # minima = skimage.morphology.h_minima(neg_dist, h_test)
        # minima_label = sp.ndimage.label(minima)[0]

        s = time.time()
        mask = fip.im_extended_min(dist, h_test)
        e = time.time()
        print("\tim_extended_min: ", e - s)
        s = time.time()
        dist2 = fip.im_impose_min(dist, mask)
        e = time.time()
        print("\tim_impose_min: ", e - s)

        lcl_min = skimage.morphology.local_minima(dist2)
        lbl = skimage.measure.label(mask)
        fip.print_numpy_image(dist2)
        # fip.print_numpy_image(dist)
        # display_mult_ims(mask, shareaxes=True)
        # display_mult_ims(lbl)
        s = time.time()
        labels = skimage.morphology.watershed(dist, lbl, mask=self.connected_mask, watershed_line=True)
        e = time.time()
        print("\twatershed: ", e - s)
        self.post_watershed_binary = sp.empty(labels.shape, dtype=bool)
        self.post_watershed_binary[labels != 0] = True
        self.post_watershed_binary = skimage.morphology.binary_erosion(self.post_watershed_binary)
        self.post_watershed_binary = skimage.morphology.remove_small_objects(self.post_watershed_binary)
        # display_mult_ims(self.post_watershed_binary)
        # print("original connected blobs: " + str(sp.amax))
        print("num blobs: " + str(sp.amax(labels)))
        self.flag_separate = True
    # END separate_connected_fibers////////////////////////////////////////////////

    # START refine_contours********************************************************
    def refine_contours(self):

        # display_mult_ims(self.membrane_image)

        height, width = self.membrane_image[:,:,0].shape
        d_p_i = 100
        height = (height / d_p_i)
        width = (width / d_p_i)
        # print(height)
        # print(width)

        fiber_labels, num_fibs = skimage.measure.label(self.fiber_mask, return_num=True)
        props = skimage.measure.regionprops(fiber_labels)
        centx = sp.empty(num_fibs)
        centy = sp.empty(num_fibs)
        bboxes = sp.empty([num_fibs, 4])
        cnt = 0
        for p in props:
            bboxes[cnt, :] = p.bbox
            centx[cnt] = p.bbox[1]
            centy[cnt] = p.bbox[0]
            cnt = cnt + 1
        # centy = -centy

        # fig, ax = plt.subplots(figsize=(7, 7))
        # ax.imshow(self.membrane_image)
        # # ax.plot(init[:, 0], init[:, 1], '--r', lw=3)
        # cnt = 0
        # for c in centx:s
        #     ax.plot(c, centy[cnt], 'ro', lw=3)
        #     cnt = cnt + 1
        # ax.set_xticks([]), ax.set_yticks([])
        # # ax.axis([0, img.shape[1], img.shape[0], 0])
        # plt.show()

        def calculate_bounds(box, tol):
            min_row = (box[0] - tol)
            max_row = (box[2] + tol)
            min_col = (box[1] - tol)
            max_col = (box[3] + tol)

            return [min_row, max_row, min_col, max_col]

        self.snakes = []
        # print(num_fibs)
        # display_mult_ims(fiber_labels)
        for blob in sp.arange(0, num_fibs):
            # print(blob)
            # im = sp.zeros(fiber_labels.shape, dtype=bool)
            # im[fiber_labels == blob] = True

            box = bboxes[blob, :].astype('int')
            tol = 20
            # shiftx = ((box[2] + tol) - (box[0] - tol))/2
            # shifty = ((box[3] + tol) - (box[1] - tol))/2
            bounds = calculate_bounds(box, tol)

            min_val = min(bounds)
            if min_val < 0:
                tol = tol + min_val
                print('new tol of ', tol, ' at blob ', blob)
                bounds = calculate_bounds(box, tol)
            region = fiber_labels[bounds[0]:bounds[1], bounds[2]:bounds[3]]
            orig = self.membrane_image[box[0] - tol:box[2] + tol, box[1] - tol:box[3] + tol, :]
            seed_region = sp.zeros(region.shape, dtype=bool)
            seed_region[region == blob + 1] = True

            asdf = seed_region.astype('uint8')
            img, contours, heirarchy = cv.findContours(asdf, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
            x = contours[0][:, :, 0]
            x = x.flatten()
            y = contours[0][:, :, 1]
            y = y.flatten()
            init = sp.array([x, y]).T
            # plt.plot(contours[0][:, :, 0], -contours[0][:, :, 1])
            # plt.show()

            gray_mem_region = skimage.color.rgb2gray(orig)
            snake = skimage.segmentation.active_contour(orig,
                                                        init,
                                                        bc='fixed')
            # print(snake[:, 0])
            tmp_snake = snake
            # print(centx[blob])
            tmp_snake[:, 0] = tmp_snake[:, 0] + centx[blob] - tol#- shiftx
            tmp_snake[:, 1] = tmp_snake[:, 1] + centy[blob] - tol#- shifty
            # print(tmp_snake)
            self.snakes.append(tmp_snake)
            # print(type(snake))
            # print(snake.shape)
            # fig, ax = plt.subplots(figsize=(7, 7))
            # ax.imshow(orig)
            # ax.plot(init[:, 0], init[:, 1], '--r', lw=3)
            # ax.plot(snake[:, 0], snake[:, 1], '-b', lw=3)
            # ax.set_xticks([]), ax.set_yticks([])
            # ax.axis([0, img.shape[1], img.shape[0], 0])
            # plt.show()
            # display_mult_ims(im)


        # fig, ax = plt.subplots(figsize=(width, height), dpi=d_p_i)  # figsize=(7, 7)
        # ax.imshow(self.membrane_image)
        # # ax.plot(init[:, 0], init[:, 1], '--r', lw=3)
        # # for snk in self.snakes:
        #     # print(snk.shape)
        #     # print(type(snk))
        #     # ax.plot(snk[:, 0], snk[:, 1], '--r', lw=1)
        # ax.set_xticks([]), ax.set_yticks([])
        # # ax.axis([0, img.shape[1], img.shape[0], 0])
        # plt.show()
        # fig.savefig("throwaway.png", bbox_inches='tight')


        # arbitrarily choose blob x for testing
        # test_blob_num = 1237
        # box = self.blob_metrics.bboxes[test_blob_num,:].astype('int')
        # tol = 1
        # region = self.blob_labels[box[0]-tol:box[2]+tol,box[1]-tol:box[3]+tol]
        # orig = self.membrane_image[box[0]-tol:box[2]+tol,box[1]-tol:box[3]+tol,:]
        # seed_region = sp.zeros(region.shape, dtype=bool)
        # seed_region[region == test_blob_num+1] = True

        # print(seed_region.dtype)
        # im = im_overlay(seed_region, orig)
        # display_mult_ims(im)

        # eroded = skimage.morphology.binary_erosion(seed_region)
        # display_mult_ims(seed_region, eroded)
        # perim makes the points for the active contour snake
        # perim = sp.logical_xor(seed_region, eroded)

        # asdf = seed_region.astype('uint8')
        # img, contours, heirarchy = cv.findContours(asdf, cv.RETR_TREE, cv.CHAIN_APPROX_SIMPLE)
        # x = contours[0][:, :, 0]
        # x = x.flatten()
        # y = contours[0][:, :, 1]
        # y = y.flatten()
        # init = sp.array([x, y]).T
        # plt.plot(contours[0][:, :, 0], -contours[0][:, :, 1])
        # plt.show()
        #
        # gray_mem_region = skimage.color.rgb2gray(orig)
        # snake = skimage.segmentation.active_contour(skimage.filters.gaussian(gray_mem_region),
        #                                             init,
        #                                             bc='periodic')

        # fig, ax = plt.subplots(figsize=(7, 7))
        # ax.imshow(gray_mem_region, cmap=plt.cm.gray)
        # ax.plot(init[:, 0], init[:, 1], '--r', lw=3)
        # ax.plot(snake[:, 0], snake[:, 1], '-b', lw=3)
        # ax.set_xticks([]), ax.set_yticks([])
        # ax.axis([0, img.shape[1], img.shape[0], 0])
        # plt.show()

    # END refine_contours//////////////////////////////////////////////////////////



    # START overlay_blobs**********************************************************
    # Description:
    #     Given a PIL image of the stain's fibers_binary, this function overlays it
    #     with red numbers corresponding to the blob number
    # Inputs:
    #     pil_im: a PIL image of this stain's fibers_binary data
    # Returns:
    #     pil_im: a PIL image of blob numbers overlayed on the stain's fibers_binary
    def overlay_blobs(self, pil_im):
        fnt = ImageFont.truetype("arial.ttf", 20)
        color = 'rgb(255, 0, 0)'  # red color

        draw = ImageDraw.Draw(pil_im)
        for x in range(self.num_blobs):
            cent_num = str(x)
            draw.text((self.blob_metrics.centroidsx[x], self.blob_metrics.centroidsy[x]), cent_num, fill=color,
                      font=fnt)

        return pil_im

    # END overlay_blobs////////////////////////////////////////////////////////////

    # START interactive_blobs******************************************************
    def interactive_blobs(self):
        blob_selector(self.blob_labels, self.membrane_image)
    # END interactive_blobs////////////////////////////////////////////////////////

    # START create_classification_image********************************************
    def create_classification_image(self):
        overlayed = copy.deepcopy(self.membrane_image)
        overlayed[self.fiber_mask, 0] = 1
        overlayed[self.fiber_mask, 1] = 1
        overlayed[self.fiber_mask, 2] = 0

        overlayed[self.connected_mask, 0] = 0
        overlayed[self.connected_mask, 1] = 1
        overlayed[self.connected_mask, 2] = 1

        overlayed[self.space_mask, 0] = 1
        overlayed[self.space_mask, 1] = 0
        overlayed[self.space_mask, 2] = 1

        return overlayed
    # END create_classification_image//////////////////////////////////////////////

    # START display_classifications************************************************
    # Description:
    #     Displays the stain with colored classifications of found blobs:
    #     Yellow = Fiber
    #     Cyan = Connected Fiber
    #     Purple = Space
    # Inputs:
    #     overlayed: A copy of the masked stain
    # Returns:
    #     N/A
    def display_classifications(self):
        overlayed = self.create_classification_image()

        display_mult_ims(overlayed)
    # END display_classifications//////////////////////////////////////////////////

    # START display_contours*******************************************************
    def display_contours(self):
        fig, ax = plt.subplots()
        ax.imshow(self.membrane_image)
        # ax.plot(init[:, 0], init[:, 1], '--r', lw=3)
        for snk in self.snakes:
            ax.plot(snk[:, 0], snk[:, 1], '--r', lw=1)
        ax.set_xticks([]), ax.set_yticks([])
        # ax.axis([0, img.shape[1], img.shape[0], 0])
        plt.show()
        # fig.canvas.draw()
        # plt.savefig('test_im_qual.eps', format='eps', dpi=1000)
        data = sp.fromstring(fig.canvas.tostring_rgb(), dtype=sp.uint8, sep='')
        data = data.reshape(fig.canvas.get_width_height()[::-1] + (3,))
        pil_im = fip.numpy_rgb_to_pil(data)
        self.contour_image = pil_im
    # END display_contours/////////////////////////////////////////////////////////



    # START block_proc*************************************************************
    # Description:
    #     Use to break a MuscleStain process that is performed on an image into smaller
    #     chunks to use less memory.
    # Inputs:
    #     im_gray: Image to process
    #     block_size: The shape of the blocks to be processed individually specified by
    #                 a tuple of 2 numbers
    #     func: A function call to the function you want performed on each block
    # Returns:
    #     stitched: The processed image
    def block_proc(self, im_gray, block_size, func):
        stitched = sp.empty(im_gray.shape)

        bs = block_size
        x, y = im_gray.shape
        # Calculate size of x chunks
        x_chunks = math.ceil(x / block_size[1])
        x_tol = math.ceil(block_size[1] * 0.05)
        x_remain = x % block_size[1]
        if (x_remain == 0): x_remain = block_size[1]
        # Calculate size of y chunks
        y_chunks = math.ceil(y / block_size[0])
        y_tol = math.ceil(block_size[0] * 0.05)
        y_remain = y % block_size[0]
        if (y_remain == 0): y_remain = block_size[0]

        for nx in range(x_chunks):
            for ny in range(y_chunks):
                if (nx != x_chunks-1):
                    if (ny != y_chunks-1):  # Both right
                        print("X- ", nx * bs[1], ":", (nx * bs[1]) + bs[1] - 1)
                        print("Y- ", ny * bs[0], ":", (ny * bs[0]) + bs[0] - 1)
                        temp = func(
                            im_gray[nx * bs[1]:(nx * bs[1]) + bs[1] + x_tol - 1, ny * bs[0]:(ny * bs[0]) + bs[0] + y_tol - 1],
                            scale_range=(1, 2),
                            scale_step=0.25,
                            beta1=1,
                            beta2=10)
                        stitched[nx * bs[1]:(nx * bs[1]) + bs[1], ny * bs[0]:(ny * bs[0]) + bs[0]] = (
                            temp[0:bs[1], 0:bs[0]]
                        )

                    else:  # X right, Y wrong
                        print("X- ", nx * bs[1], ":", (nx * bs[1]) + bs[1] - 1)
                        print("Y- ", ny * bs[0], ":", (ny * bs[0]) + y_remain - 1)
                        temp = func(
                            im_gray[nx * bs[1]:(nx * bs[1]) + x_tol + bs[1] - 1, ny * bs[0]:(ny * bs[0]) + y_remain],
                            scale_range=(1, 2),
                            scale_step=0.25,
                            beta1=1,
                            beta2=10)
                        stitched[nx * bs[1]:(nx * bs[1]) + bs[1], ny * bs[0]:(ny * bs[0]) + y_remain] = (
                            temp[0:bs[1], 0:y_remain]
                        )
                else:
                    if (ny != y_chunks-1):  # X wrong, Y right
                        print("X- ", nx * bs[1], ":", (nx * bs[1]) + x_remain - 1)
                        print("Y- ", ny * bs[0], ":", (ny * bs[0]) + bs[0] - 1)
                        temp = func(
                            im_gray[nx * bs[1]:(nx * bs[1]) + x_remain, ny * bs[0]:(ny * bs[0]) + y_tol + bs[0] - 1],
                            scale_range=(1, 2),
                            scale_step=0.25,
                            beta1=1,
                            beta2=10)
                        stitched[nx * bs[1]:(nx * bs[1]) + x_remain, ny * bs[0]:(ny * bs[0]) + bs[0]] = (
                            temp[0:x_remain, 0:bs[0]]
                        )
                    else:  # X wrong, Y wrong
                        print("X- ", nx * bs[1], ":", (nx * bs[1]) + x_remain - 1)
                        print("Y- ", ny * bs[0], ":", (ny * bs[0]) + y_remain - 1)
                        temp = func(
                            im_gray[nx * bs[1]:(nx * bs[1]) + x_remain, ny * bs[0]:(ny * bs[0]) + y_remain],
                            scale_range=(1, 2),
                            scale_step=0.25,
                            beta1=1,
                            beta2=10)
                        stitched[nx * bs[1]:(nx * bs[1]) + x_remain, ny * bs[0]:(ny * bs[0]) + y_remain] = (
                            temp[0:x_remain, 0:y_remain]
                        )

        return stitched
    # END block_proc///////////////////////////////////////////////////////////////

    # START write_data*************************************************************
    def write_data(self):
        wb = openpyxl.load_workbook(self.excel_book)
        ws = wb['Initial Blob Image']
        try:
            self.initial_blob_image.save('temp_im.png')
            # display_mult_ims(self.initial_blob_image)
            img = openpyxl.drawing.image.Image('temp_im.png')
            ws.add_image(img, 'A5')
        except AttributeError:
            pass

        ws = wb['Initial Blob Classifications']
        try:
            self.mask_image_1.save('temp_im1.png')
            # display_mult_ims(self.initial_blob_image)
            img = openpyxl.drawing.image.Image('temp_im1.png')
            ws.add_image(img, 'A5')
        except AttributeError:
            pass

        ws2 = wb['Post-Watershed Image']
        try:
            self.post_watershed_image.save('temp_im2.png')
            # display_mult_ims(self.post_watershed_image)
            img = openpyxl.drawing.image.Image('temp_im2.png')
            ws2.add_image(img, 'A5')
        except AttributeError:
            pass

        ws = wb['Post-Watershed Classifications']
        try:
            self.mask_image_2.save('temp_im3.png')
            # display_mult_ims(self.initial_blob_image)
            img = openpyxl.drawing.image.Image('temp_im3.png')
            ws.add_image(img, 'A5')
        except AttributeError:
            pass

        # ws = wb['Contours']
        # self.contour_image.save('temp_im4.png')
        # # display_mult_ims(self.initial_blob_image)
        # img = openpyxl.drawing.image.Image('temp_im4.png')
        # ws.add_image(img, 'A5')

        for sheet_name in self.image_sheet_names:
            ws = wb[sheet_name]

        wb.save(self.excel_book)

        try:
            compressed_filename = copy.deepcopy(self.excel_book)
            compressed_filepath = compressed_filename.replace('.xlsx', '.zip')
            with zipfile.ZipFile(compressed_filepath, 'w') as zipper:
                zipper.write('temp_im.png')
                zipper.write('temp_im1.png')
                zipper.write('temp_im2.png')
                zipper.write('temp_im3.png')
                zipper.write(self.excel_book, arcname=os.path.basename(self.excel_book))
                zipper.close()
        except FileNotFoundError:
            zipper.close()
            print("Not all images included in zip file. Likely because not all processing was done on stain before writing data")

    # END write_data///////////////////////////////////////////////////////////////